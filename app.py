#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
(주)한영피엔에스 IMS 문서 미리보기 웹앱 v2.2
변경사항:
  - PORT 환경변수 지원 (Render.com 등 클라우드 배포 대응)
  - DOCS_ROOT 환경변수 지원 (배포 환경 자동 적응)
  - 상대경로 우선 탐색: ./docs → 환경변수 → 기본 경로 순
"""

from flask import Flask, jsonify, render_template_string, request, send_file, Response
from docx import Document
import openpyxl
import os, json, html, re, io, zipfile, urllib.parse

try:
    import segno  # 순수 파이썬 QR 코드 생성 (의존성 없음)
    _HAS_SEGNO = True
except ImportError:
    _HAS_SEGNO = False

app = Flask(__name__)

# ─── 캐시 (메모리) ──────────────────────────────────────
_preview_cache = {}

# ─── DOCS_ROOT: 환경변수 > 상대경로 > 절대경로 순 탐색 ──
def _find_docs_root():
    # 1) 환경변수 우선
    env_path = os.environ.get('DOCS_ROOT', '')
    if env_path and os.path.isdir(env_path):
        return env_path
    # 2) 앱 파일 기준 상대경로 ./docs
    base = os.path.dirname(os.path.abspath(__file__))
    rel = os.path.join(base, 'docs')
    if os.path.isdir(rel):
        return rel
    # 3) 로컬 개발 환경 절대경로
    local = '/home/user/ims_full/output_final'
    if os.path.isdir(local):
        return local
    # 4) 마지막 fallback
    return rel  # 없어도 경로는 반환 (에러는 API에서 처리)

DOCS_ROOT = _find_docs_root()
print(f'📁 DOCS_ROOT: {DOCS_ROOT}  (존재: {os.path.isdir(DOCS_ROOT)})')

STANDARD_INFO = {
    'ISO9001':  {'name': 'ISO 9001 품질경영',    'color': '#1F3864', 'icon': '🏆', 'badge': 'QMS'},
    'ISO14001': {'name': 'ISO 14001 환경경영',   'color': '#375623', 'icon': '🌿', 'badge': 'EMS'},
    'ISO45001': {'name': 'ISO 45001 안전보건',   'color': '#843C0C', 'icon': '🦺', 'badge': 'OH&S'},
    'ISO22000': {'name': 'ISO 22000 식품안전',   'color': '#7B4C00', 'icon': '🍽', 'badge': 'FSMS'},
    'ISO37001': {'name': 'ISO 37001 반부패',     'color': '#4B0082', 'icon': '⚖', 'badge': 'ABMS'},
    'ISO37301': {'name': 'ISO 37301 준법경영',   'color': '#1A3C5C', 'icon': '📋', 'badge': 'CMS'},
    'ISO27001': {'name': 'ISO 27001 정보보안',   'color': '#7B0000', 'icon': '🔒', 'badge': 'ISMS'},
    'IMS':      {'name': 'IMS 통합경영시스템',   'color': '#0A2342', 'icon': '🔗', 'badge': 'IMS'},
}

# SVG 파비콘
FAVICON_SVG = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32">
  <rect width="32" height="32" rx="6" fill="#1F3864"/>
  <text x="16" y="22" font-size="18" text-anchor="middle" fill="white" font-family="Arial">P</text>
</svg>'''

@app.route('/favicon.ico')
@app.route('/favicon.svg')
def favicon():
    return Response(FAVICON_SVG, mimetype='image/svg+xml')


def get_doc_type(filename):
    fn = filename.upper()
    if '-M-' in fn or 'MANUAL' in fn or '매뉴얼' in fn:   return '매뉴얼', '#1F3864'
    if '-P-' in fn or 'PROC' in fn or '절차서' in fn:     return '절차서', '#2E75B6'
    if '-G-' in fn or 'GUIDE' in fn or '지침서' in fn:    return '지침서', '#70AD47'
    if '-F-' in fn or 'FORM' in fn or '양식' in fn:       return '양식', '#ED7D31'
    if 'KPI' in fn or 'CAL' in fn or 'AUDIT' in fn or \
       'CERT' in fn or 'MAP' in fn or 'RISK' in fn:       return '통합관리', '#7030A0'
    return '문서', '#595959'


# ─────────────────────────────────────────────────────
# 📷 바코드 / QR 코드 조회 기능
# ─────────────────────────────────────────────────────
def extract_doc_code(filename):
    """파일명에서 문서코드(첫 '_' 앞 부분) 추출. 예: QMS-G-001"""
    base = filename.rsplit('.', 1)[0]
    return base.split('_', 1)[0].strip().upper()


def build_doc_index():
    """전체 문서를 스캔하여 조회용 인덱스 생성"""
    index = []
    for std_key, info in STANDARD_INFO.items():
        std_path = os.path.join(DOCS_ROOT, std_key)
        if not os.path.isdir(std_path):
            continue
        for fname in sorted(os.listdir(std_path)):
            if not (fname.endswith('.docx') or fname.endswith('.xlsx')):
                continue
            dtype, dcolor = get_doc_type(fname)
            index.append({
                'std': std_key,
                'std_name': info['name'],
                'code': extract_doc_code(fname),
                'name': fname,
                'display': fname.replace('.docx', '').replace('.xlsx', ''),
                'type': dtype,
                'color': dcolor,
                'ext': fname.rsplit('.', 1)[-1].upper(),
            })
    return index


def normalize_scan(raw):
    """스캔된 원본 값을 조회 키로 정규화.
    - URL 이면 code/q/d 파라미터 또는 마지막 경로 세그먼트 추출
    - 그 외에는 원본 텍스트 사용
    """
    raw = (raw or '').strip()
    if not raw:
        return ''
    if '://' in raw or raw.startswith('/'):
        try:
            parsed = urllib.parse.urlparse(raw)
            qs = urllib.parse.parse_qs(parsed.query)
            for key in ('code', 'q', 'd'):
                if qs.get(key):
                    return qs[key][0].strip()
            segs = [s for s in parsed.path.split('/') if s]
            if segs:
                return urllib.parse.unquote(segs[-1]).strip()
        except Exception:
            pass
    return raw


def lookup_code(raw):
    """스캔/입력된 코드를 해석하여 관련 문서 정보를 반환"""
    code = normalize_scan(raw)
    cl = code.lower()
    index = build_doc_index()

    if not cl:
        return {'kind': 'empty', 'query': code, 'results': []}

    # 1) 규격 키 / 배지 매칭 (예: ISO9001, QMS, IMS)
    for std_key, info in STANDARD_INFO.items():
        if cl == std_key.lower() or cl == info['badge'].lower():
            files = [d for d in index if d['std'] == std_key]
            std = {'key': std_key, 'name': info['name'], 'icon': info['icon'],
                   'color': info['color'], 'badge': info['badge'], 'count': len(files)}
            return {'kind': 'standard', 'query': code, 'standard': std, 'results': files}

    # 2) 문서코드 정확 매칭 (예: QMS-G-001)
    matches = [d for d in index if d['code'].lower() == cl]
    if matches:
        return {'kind': 'document', 'query': code, 'results': matches}

    # 3) 파일명 / 코드 부분 일치
    matches = [d for d in index if cl in d['name'].lower() or cl in d['code'].lower()]
    if matches:
        return {'kind': 'search', 'query': code, 'results': matches[:40]}

    # 4) 표시명 / 규격명 키워드 검색
    matches = [d for d in index
               if cl in d['display'].lower() or cl in d['std_name'].lower()]
    return {'kind': 'search' if matches else 'none', 'query': code, 'results': matches[:40]}


def docx_to_html(path):
    """Word 문서를 HTML로 변환 (개선 버전)"""
    try:
        doc = Document(path)
        NS = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
        html_parts = ['<div class="doc-content">']
        in_list = False

        for element in doc.element.body:
            tag = element.tag.split('}')[-1] if '}' in element.tag else element.tag

            if tag == 'p':
                # 스타일명 추출
                style_name = ''
                try:
                    pPr = element.find(f'.//{{{NS}}}pStyle')
                    if pPr is not None:
                        style_name = pPr.get(f'{{{NS}}}val', '')
                except:
                    pass

                # 텍스트 수집 (굵게/이탤릭 포함)
                text_parts = []
                for r in element.findall(f'.//{{{NS}}}r'):
                    t = r.find(f'{{{NS}}}t')
                    if t is not None and t.text:
                        rPr = r.find(f'{{{NS}}}rPr')
                        bold = rPr is not None and rPr.find(f'{{{NS}}}b') is not None
                        italic = rPr is not None and rPr.find(f'{{{NS}}}i') is not None
                        t_text = html.escape(t.text)
                        if bold and italic:
                            t_text = f'<em><strong>{t_text}</strong></em>'
                        elif bold:
                            t_text = f'<strong>{t_text}</strong>'
                        elif italic:
                            t_text = f'<em>{t_text}</em>'
                        text_parts.append(t_text)

                text = ''.join(text_parts).strip()

                # 스타일 → HTML 태그 매핑 (Heading 1/2/3 + 한글 변형 모두 처리)
                sl = style_name.lower().replace(' ', '').replace('-', '')
                is_h1 = sl in ('heading1', '1', 'title') or style_name == 'Heading 1'
                is_h2 = sl in ('heading2', '2', 'subtitle') or style_name == 'Heading 2'
                is_h3 = sl in ('heading3', '3') or style_name == 'Heading 3'
                is_list = sl in ('listbullet', 'listbullet2', 'listparagraph') or style_name == 'List Bullet'

                if is_list:
                    if not in_list:
                        html_parts.append('<ul class="doc-ul">')
                        in_list = True
                    if text:
                        html_parts.append(f'<li class="doc-li">{text}</li>')
                    continue

                # 리스트 닫기
                if in_list:
                    html_parts.append('</ul>')
                    in_list = False

                if not text:
                    html_parts.append('<p class="empty-p">&nbsp;</p>')
                    continue

                if is_h1:
                    html_parts.append(f'<h2 class="doc-h1">{text}</h2>')
                elif is_h2:
                    html_parts.append(f'<h3 class="doc-h2">{text}</h3>')
                elif is_h3:
                    html_parts.append(f'<h4 class="doc-h3">{text}</h4>')
                else:
                    html_parts.append(f'<p class="doc-p">{text}</p>')

            elif tag == 'tbl':
                # 리스트 닫기
                if in_list:
                    html_parts.append('</ul>')
                    in_list = False

                html_parts.append('<div class="table-wrap"><table class="doc-table">')
                rows = element.findall(f'.//{{{NS}}}tr')

                for i, row in enumerate(rows):
                    html_parts.append('<tr>')
                    cells = row.findall(f'{{{NS}}}tc')
                    for cell in cells:
                        # gridSpan (열 병합) 처리
                        colspan = ''
                        grid_span = cell.find(f'.//{{{NS}}}gridSpan')
                        if grid_span is not None:
                            span_val = grid_span.get(f'{{{NS}}}val', '1')
                            if span_val and int(span_val) > 1:
                                colspan = f' colspan="{span_val}"'

                        texts = []
                        for p in cell.findall(f'.//{{{NS}}}p'):
                            cell_texts = []
                            for r in p.findall(f'.//{{{NS}}}r'):
                                t = r.find(f'{{{NS}}}t')
                                if t is not None and t.text:
                                    rPr = r.find(f'{{{NS}}}rPr')
                                    bold = rPr is not None and rPr.find(f'{{{NS}}}b') is not None
                                    ct = html.escape(t.text)
                                    cell_texts.append(f'<strong>{ct}</strong>' if bold else ct)
                            if cell_texts:
                                texts.append(''.join(cell_texts))
                        cell_content = '<br>'.join(texts) if texts else '&nbsp;'
                        tag_t = 'th' if i == 0 else 'td'
                        html_parts.append(f'<{tag_t}{colspan}>{cell_content}</{tag_t}>')
                    html_parts.append('</tr>')
                html_parts.append('</table></div>')

        if in_list:
            html_parts.append('</ul>')
        html_parts.append('</div>')
        return '\n'.join(html_parts)

    except Exception as e:
        return f'<div class="error-box">⚠️ 문서 변환 오류: {html.escape(str(e))}</div>'


def xlsx_to_html(path):
    """Excel 파일을 HTML로 변환 (개선 버전)"""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames

        html_parts = ['<div class="xlsx-content">']
        # 시트 탭
        html_parts.append('<div class="sheet-tabs" id="sheetTabsBar">')
        for i, sname in enumerate(sheets):
            active = 'active' if i == 0 else ''
            safe_name = html.escape(sname)
            html_parts.append(
                f'<button class="sheet-tab {active}" onclick="showSheet({i})" id="tab-{i}">'
                f'{safe_name}</button>'
            )
        html_parts.append('</div>')

        for sheet_idx, sname in enumerate(sheets):
            ws = wb[sname]
            display = 'block' if sheet_idx == 0 else 'none'
            html_parts.append(
                f'<div class="sheet-content" id="sheet-{sheet_idx}" style="display:{display}">'
            )
            html_parts.append(f'<div class="sheet-title">📊 {html.escape(sname)}</div>')
            html_parts.append('<div class="table-wrap"><table class="doc-table xlsx-table">')

            rows_data = []
            for row in ws.iter_rows(max_row=120, values_only=True):
                rows_data.append(row)

            # 실제 첫 번째 데이터 행을 헤더로 사용
            header_done = False
            for r_idx, row in enumerate(rows_data):
                # 완전히 빈 행 스킵
                if all(v is None or str(v).strip() == '' for v in row):
                    continue
                html_parts.append('<tr>')
                for val in row:
                    v = '&nbsp;' if val is None else html.escape(str(val))
                    if not header_done:
                        html_parts.append(f'<th>{v}</th>')
                    else:
                        html_parts.append(f'<td>{v}</td>')
                html_parts.append('</tr>')
                header_done = True

            html_parts.append('</table></div>')
            total_rows = ws.max_row or 0
            if total_rows > 120:
                html_parts.append(f'<p class="limit-note">※ 처음 120행만 표시 (전체 약 {total_rows}행)</p>')
            html_parts.append('</div>')

        wb.close()
        html_parts.append('</div>')
        return '\n'.join(html_parts)

    except Exception as e:
        return f'<div class="error-box">⚠️ Excel 변환 오류: {html.escape(str(e))}</div>'


# ─────────────────────────────────────────────────────
# API Routes
# ─────────────────────────────────────────────────────

@app.route('/api/standards')
def api_standards():
    result = []
    for std_key, info in STANDARD_INFO.items():
        std_path = os.path.join(DOCS_ROOT, std_key)
        if os.path.exists(std_path):
            files = [f for f in os.listdir(std_path)
                     if f.endswith('.docx') or f.endswith('.xlsx')]
            result.append({
                'key': std_key,
                'name': info['name'],
                'color': info['color'],
                'icon': info['icon'],
                'badge': info['badge'],
                'count': len(files)
            })
    return jsonify(result)


@app.route('/api/files/<std>')
def api_files(std):
    std_path = os.path.join(DOCS_ROOT, std)
    if not os.path.exists(std_path):
        return jsonify({'error': '규격을 찾을 수 없습니다'}), 404

    files = []
    for fname in sorted(os.listdir(std_path)):
        if not (fname.endswith('.docx') or fname.endswith('.xlsx')):
            continue
        fpath = os.path.join(std_path, fname)
        dtype, dcolor = get_doc_type(fname)
        ext = fname.rsplit('.', 1)[-1].upper()
        files.append({
            'name': fname,
            'display': fname.replace('.docx', '').replace('.xlsx', ''),
            'type': dtype,
            'color': dcolor,
            'ext': ext,
            'size': round(os.path.getsize(fpath) / 1024, 1)
        })
    return jsonify(files)


@app.route('/api/preview/<std>/<path:filename>')
def api_preview(std, filename):
    fpath = os.path.join(DOCS_ROOT, std, filename)
    if not os.path.exists(fpath):
        return jsonify({'error': f'파일을 찾을 수 없습니다: {filename}'}), 404

    # 캐시 확인
    cache_key = f'{std}/{filename}'
    mtime = os.path.getmtime(fpath)
    if cache_key in _preview_cache:
        cached_mtime, cached_content = _preview_cache[cache_key]
        if cached_mtime == mtime:
            return jsonify({'content': cached_content, 'filename': filename, 'cached': True})

    if filename.endswith('.docx'):
        content = docx_to_html(fpath)
    elif filename.endswith('.xlsx'):
        content = xlsx_to_html(fpath)
    else:
        content = '<div class="error-box">지원하지 않는 파일 형식입니다</div>'

    # 캐시 저장
    _preview_cache[cache_key] = (mtime, content)

    return jsonify({'content': content, 'filename': filename, 'cached': False})


@app.route('/api/download/<std>/<path:filename>')
def api_download(std, filename):
    fpath = os.path.join(DOCS_ROOT, std, filename)
    if not os.path.exists(fpath):
        return '파일을 찾을 수 없습니다', 404

    # 파일명 RFC 5987 인코딩
    encoded_name = urllib.parse.quote(filename.encode('utf-8'))
    response = send_file(fpath, as_attachment=True)
    response.headers['Content-Disposition'] = (
        f"attachment; filename*=UTF-8''{encoded_name}"
    )
    return response


@app.route('/api/download-zip/<std>')
def api_download_zip(std):
    """특정 규격 전체 ZIP 다운로드"""
    std_path = os.path.join(DOCS_ROOT, std)
    if not os.path.exists(std_path):
        return '규격을 찾을 수 없습니다', 404

    std_name = STANDARD_INFO.get(std, {}).get('name', std)
    zip_name = f'한영피엔에스_{std}_{std_name.split()[0]}_문서패키지.zip'

    mem = io.BytesIO()
    with zipfile.ZipFile(mem, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname in sorted(os.listdir(std_path)):
            if fname.endswith('.docx') or fname.endswith('.xlsx'):
                fpath = os.path.join(std_path, fname)
                zf.write(fpath, fname)
    mem.seek(0)

    encoded_zip = urllib.parse.quote(zip_name.encode('utf-8'))
    response = send_file(mem, mimetype='application/zip', as_attachment=True,
                         download_name=zip_name)
    response.headers['Content-Disposition'] = (
        f"attachment; filename*=UTF-8''{encoded_zip}"
    )
    return response


@app.route('/api/stats')
def api_stats():
    total = word = excel = 0
    breakdown = {}
    for std_key in STANDARD_INFO:
        std_path = os.path.join(DOCS_ROOT, std_key)
        w = e = 0
        if os.path.exists(std_path):
            for f in os.listdir(std_path):
                if f.endswith('.docx'):   w += 1; word += 1; total += 1
                elif f.endswith('.xlsx'): e += 1; excel += 1; total += 1
        breakdown[std_key] = {'word': w, 'excel': e}
    return jsonify({'total': total, 'word': word, 'excel': excel,
                    'standards': len(STANDARD_INFO), 'breakdown': breakdown})


@app.route('/api/search')
def api_search():
    q = request.args.get('q', '').strip().lower()
    if len(q) < 2:
        return jsonify([])

    results = []
    for std_key, info in STANDARD_INFO.items():
        std_path = os.path.join(DOCS_ROOT, std_key)
        if not os.path.exists(std_path):
            continue
        for fname in sorted(os.listdir(std_path)):
            if not (fname.endswith('.docx') or fname.endswith('.xlsx')):
                continue
            if q in fname.lower() or q in info['name'].lower():
                dtype, dcolor = get_doc_type(fname)
                results.append({
                    'std': std_key,
                    'std_name': info['name'],
                    'name': fname,
                    'display': fname.replace('.docx', '').replace('.xlsx', ''),
                    'type': dtype,
                    'color': dcolor,
                    'ext': fname.rsplit('.', 1)[-1].upper()
                })
    return jsonify(results[:40])


# ─────────────────────────────────────────────────────
# 📷 바코드/QR 조회 API & QR 이미지 생성
# ─────────────────────────────────────────────────────
@app.route('/api/lookup')
def api_lookup():
    """스캔된 코드(문서코드/규격/URL/검색어)로 관련 문서 조회"""
    raw = request.args.get('code', '') or request.args.get('q', '')
    return jsonify(lookup_code(raw))


@app.route('/api/qr')
def api_qr():
    """주어진 data 문자열을 QR 코드(SVG)로 생성"""
    data = request.args.get('data', '').strip()
    if not data:
        return '조회할 data 파라미터가 필요합니다', 400
    if not _HAS_SEGNO:
        return 'QR 생성 모듈(segno)이 설치되지 않았습니다', 503
    try:
        qr = segno.make(data, error='m')
    except Exception as e:
        return f'QR 생성 실패: {e}', 400
    buff = io.BytesIO()
    qr.save(buff, kind='svg', scale=6, border=2, dark='#1F3864', light='#ffffff')
    return Response(buff.getvalue(), mimetype='image/svg+xml',
                    headers={'Cache-Control': 'public, max-age=86400'})


# ─────────────────────────────────────────────────────
# 메인 페이지 HTML
# ─────────────────────────────────────────────────────
HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>(주)한영피엔에스 IMS 문서 미리보기 시스템</title>
<link rel="icon" type="image/svg+xml" href="/favicon.svg">
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
:root {
  --brand: #1F3864;
  --brand-mid: #2E75B6;
  --brand-light: #D6E4F0;
  --sidebar-w: 290px;
  --filepanel-w: 310px;
  --header-h: 60px;
}
body { font-family: 'Malgun Gothic','맑은 고딕',sans-serif; background:#f0f2f5; color:#222; }

/* ── 헤더 ── */
.header {
  position:fixed; top:0; left:0; right:0; height:var(--header-h);
  background:var(--brand); color:#fff; display:flex; align-items:center;
  padding:0 16px; z-index:100; box-shadow:0 2px 8px rgba(0,0,0,.3); gap:14px;
}
.header-logo { font-size:17px; font-weight:700; white-space:nowrap; line-height:1.3; }
.header-logo span { font-size:11px; opacity:.7; display:block; font-weight:400; }
.header-search { flex:1; max-width:460px; position:relative; }
.header-search input {
  width:100%; padding:7px 14px 7px 36px;
  border:none; border-radius:18px; font-size:13px;
  background:rgba(255,255,255,.18); color:#fff; outline:none;
}
.header-search input::placeholder { color:rgba(255,255,255,.55); }
.header-search input:focus { background:rgba(255,255,255,.28); }
.search-icon { position:absolute; left:11px; top:50%; transform:translateY(-50%); opacity:.7; font-size:14px; }
.search-results {
  position:absolute; top:40px; left:0; right:0;
  background:#fff; border-radius:8px; box-shadow:0 8px 24px rgba(0,0,0,.18);
  max-height:340px; overflow-y:auto; z-index:200; display:none; border:1px solid #e0e0e0;
}
.search-results.show { display:block; }
.search-item {
  padding:9px 13px; cursor:pointer; display:flex; align-items:center; gap:8px;
  border-bottom:1px solid #f3f3f3; color:#222; font-size:13px;
}
.search-item:hover { background:var(--brand-light); }
.search-item-empty { padding:12px 14px; color:#aaa; font-size:13px; text-align:center; }
.s-type { font-size:10px; padding:2px 6px; border-radius:4px; color:#fff; white-space:nowrap; }
.scan-btn {
  display:flex; align-items:center; gap:5px; flex-shrink:0;
  background:rgba(255,255,255,.18); color:#fff; text-decoration:none;
  padding:7px 13px; border-radius:18px; font-size:13px; font-weight:600;
  transition:.15s; white-space:nowrap;
}
.scan-btn:hover { background:rgba(255,255,255,.30); }
.header-stats { font-size:11px; opacity:.85; white-space:nowrap; display:flex; gap:10px; }
.stat-item { text-align:center; }
.stat-item strong { display:block; font-size:17px; font-weight:800; }

/* ── 레이아웃 ── */
.layout { display:flex; margin-top:var(--header-h); height:calc(100vh - var(--header-h)); overflow:hidden; }

/* ── 사이드바 ── */
.sidebar {
  width:var(--sidebar-w); background:#fff; overflow-y:auto;
  border-right:1px solid #e0e0e0; flex-shrink:0;
}
.sidebar-title {
  padding:10px 14px 6px; font-size:10px; color:#999;
  text-transform:uppercase; font-weight:700; letter-spacing:.6px;
  border-bottom:1px solid #f0f0f0;
}
.std-card {
  display:flex; align-items:center; gap:10px;
  padding:10px 14px; cursor:pointer; transition:.12s;
  border-left:3px solid transparent;
}
.std-card:hover { background:#f8f9fb; }
.std-card.active { background:var(--brand-light); border-left-color:var(--brand); }
.std-icon { font-size:18px; width:26px; text-align:center; }
.std-info { flex:1; min-width:0; }
.std-name { font-size:12px; font-weight:700; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.std-meta { font-size:10px; color:#999; margin-top:1px; }
.std-badge {
  font-size:9px; padding:2px 6px; border-radius:4px; color:#fff;
  font-weight:700; white-space:nowrap; letter-spacing:.3px;
}

/* ── 파일 목록 패널 ── */
.file-panel {
  width:var(--filepanel-w); background:#fafafa; overflow-y:auto;
  border-right:1px solid #e0e0e0; flex-shrink:0; display:flex; flex-direction:column;
}
.file-panel-header {
  position:sticky; top:0; background:#fff; padding:12px 14px;
  border-bottom:1px solid #e8e8e8; z-index:10; flex-shrink:0;
}
.file-panel-title { font-size:14px; font-weight:700; color:var(--brand); }
.file-panel-count { font-size:11px; color:#999; margin-top:1px; }
.file-filter { display:flex; gap:4px; margin-top:8px; flex-wrap:wrap; }
.filter-btn {
  padding:3px 9px; border-radius:11px; border:1px solid #ddd;
  font-size:10px; cursor:pointer; background:#fff; color:#555; transition:.12s;
  font-family: inherit;
}
.filter-btn.active { background:var(--brand); color:#fff; border-color:var(--brand); }
.filter-btn:hover:not(.active) { background:#f0f0f0; }
.file-list { padding:6px; flex:1; }
.file-item {
  display:flex; align-items:center; gap:7px;
  padding:7px 9px; border-radius:6px; cursor:pointer;
  margin-bottom:2px; transition:.12s;
}
.file-item:hover { background:#eff2f7; }
.file-item.active { background:var(--brand-light); box-shadow:inset 0 0 0 1px var(--brand-mid); }
.file-ext {
  font-size:9px; padding:2px 5px; border-radius:3px;
  font-weight:800; color:#fff; min-width:34px; text-align:center; flex-shrink:0;
}
.file-ext.docx { background:#2E75B6; }
.file-ext.xlsx { background:#217346; }
.file-name { font-size:11px; flex:1; min-width:0; }
.file-name strong { display:block; font-weight:600; color:#222; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.file-name small { color:#aaa; font-size:10px; }
.file-type-badge { font-size:9px; padding:1px 5px; border-radius:3px; color:#fff; flex-shrink:0; }

/* ── 미리보기 패널 ── */
.preview-panel { flex:1; overflow:hidden; display:flex; flex-direction:column; min-width:0; }
.preview-header {
  background:#fff; padding:12px 18px; border-bottom:1px solid #e8e8e8;
  display:flex; align-items:center; gap:10px; flex-shrink:0;
}
.preview-filename { flex:1; min-width:0; }
.preview-filename h3 { font-size:14px; color:var(--brand); font-weight:700; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.preview-filename p { font-size:11px; color:#999; margin-top:1px; }
.btn-group { display:flex; gap:6px; flex-shrink:0; }
.btn-download {
  padding:6px 13px; background:var(--brand); color:#fff;
  border:none; border-radius:5px; cursor:pointer; font-size:12px;
  display:flex; align-items:center; gap:5px; white-space:nowrap;
  text-decoration:none; font-family:inherit;
}
.btn-download:hover { background:#162a4d; }
.btn-zip {
  padding:6px 13px; background:#217346; color:#fff;
  border:none; border-radius:5px; cursor:pointer; font-size:12px;
  display:flex; align-items:center; gap:5px; white-space:nowrap;
  text-decoration:none; font-family:inherit;
}
.btn-zip:hover { background:#155230; }
.preview-body { flex:1; overflow-y:auto; padding:20px; }

/* ── 문서 내용 스타일 ── */
.doc-content { max-width:860px; margin:0 auto; }
.doc-h1 { font-size:15px; color:#1F3864; border-bottom:2px solid #1F3864; padding-bottom:5px; margin:18px 0 9px; }
.doc-h2 { font-size:13px; color:#2E75B6; margin:14px 0 7px; font-weight:700; }
.doc-h3 { font-size:12px; color:#70AD47; margin:10px 0 5px; font-weight:700; }
.doc-p { font-size:12px; line-height:1.75; margin:3px 0; color:#333; }
.doc-ul { margin:4px 0 4px 16px; padding:0; }
.doc-li { font-size:12px; line-height:1.75; margin:1px 0; color:#444; list-style:disc; }
.empty-p { height:5px; }
.table-wrap { overflow-x:auto; margin:10px 0; border-radius:4px; box-shadow:0 1px 3px rgba(0,0,0,.08); }
.doc-table { border-collapse:collapse; width:100%; font-size:11px; }
.doc-table th { background:#1F3864; color:#fff; padding:7px 9px; text-align:left; font-weight:700; border:1px solid #1a305a; }
.doc-table td { padding:6px 9px; border:1px solid #dde3ea; vertical-align:top; }
.doc-table tr:nth-child(even) td { background:#f8fafc; }
.doc-table tr:hover td { background:#ebf3fb; transition:.1s; }
.error-box { background:#fff3cd; border:1px solid #ffc107; padding:12px 16px; border-radius:6px; color:#856404; font-size:13px; }

/* ── Excel 스타일 ── */
.xlsx-content { max-width:100%; }
.sheet-tabs { display:flex; flex-wrap:wrap; gap:4px; padding:10px 12px; background:#f5f5f5; border-bottom:1px solid #ddd; position:sticky; top:0; z-index:10; }
.sheet-tab { padding:4px 10px; border-radius:4px; border:1px solid #ccc; background:#fff; cursor:pointer; font-size:11px; transition:.12s; font-family:inherit; }
.sheet-tab.active { background:#217346; color:#fff; border-color:#217346; }
.sheet-tab:hover:not(.active) { background:#d5e8d4; border-color:#aaa; }
.sheet-title { font-size:12px; font-weight:700; color:#217346; padding:8px 12px 4px; }
.xlsx-table th { background:#217346; font-size:11px; text-align:center; }
.xlsx-table td { font-size:11px; }
.limit-note { font-size:10px; color:#aaa; padding:6px 12px; font-style:italic; }

/* ── 웰컴 화면 ── */
.welcome-screen {
  display:flex; flex-direction:column; align-items:center; justify-content:center;
  height:100%; color:#bbb; text-align:center; padding:40px;
}
.welcome-screen h2 { font-size:20px; color:#ccc; margin-bottom:10px; }
.welcome-screen p { font-size:13px; line-height:1.9; }

/* ── 로딩 ── */
.loading { display:flex; align-items:center; justify-content:center; height:160px; flex-direction:column; gap:12px; }
.spinner { width:32px; height:32px; border:3px solid #e8e8e8; border-top-color:var(--brand); border-radius:50%; animation:spin .7s linear infinite; }
@keyframes spin { to { transform:rotate(360deg); } }
.loading-text { font-size:12px; color:#aaa; }

/* ── 빈 상태 ── */
.empty-state { text-align:center; padding:32px 16px; color:#ccc; font-size:13px; }

/* ── 토스트 알림 ── */
.toast {
  position:fixed; bottom:24px; right:24px; background:#333; color:#fff;
  padding:10px 18px; border-radius:6px; font-size:13px; z-index:9999;
  opacity:0; transition:opacity .3s; pointer-events:none;
}
.toast.show { opacity:1; }

/* ── 파일 정보 카드 ── */
.file-info-card {
  display:flex; align-items:center; gap:10px; margin-bottom:16px; padding:14px 16px;
  background:#fff; border-radius:8px; box-shadow:0 1px 5px rgba(0,0,0,.08);
}
</style>
</head>
<body>

<!-- 헤더 -->
<header class="header">
  <div class="header-logo">
    (주)한영피엔에스
    <span>IMS 통합경영시스템 문서 미리보기</span>
  </div>
  <div class="header-search">
    <span class="search-icon">🔍</span>
    <input type="text" id="searchInput" placeholder="문서명 검색... (2자 이상)" autocomplete="off">
    <div class="search-results" id="searchResults"></div>
  </div>
  <a href="/scan" class="scan-btn" title="바코드/QR 스캔">📷 <span>스캔</span></a>
  <div class="header-stats" id="headerStats">
    <div class="stat-item"><strong id="statTotal">-</strong>총 문서</div>
    <div class="stat-item"><strong id="statWord">-</strong>Word</div>
    <div class="stat-item"><strong id="statExcel">-</strong>Excel</div>
  </div>
</header>

<!-- 메인 레이아웃 -->
<div class="layout">
  <!-- 좌측: 규격 목록 -->
  <aside class="sidebar">
    <div class="sidebar-title">📌 ISO 규격 선택</div>
    <div id="stdList"></div>
  </aside>

  <!-- 중간: 파일 목록 -->
  <div class="file-panel" id="filePanel">
    <div class="welcome-screen" style="height:100%">
      <div style="font-size:44px;margin-bottom:10px">📂</div>
      <h2>규격을 선택하세요</h2>
      <p>좌측에서 ISO 규격을 클릭하면<br>해당 문서 목록이 표시됩니다</p>
    </div>
  </div>

  <!-- 우측: 미리보기 -->
  <div class="preview-panel">
    <div id="previewHeaderBar" style="display:none" class="preview-header">
      <div class="preview-filename">
        <h3 id="previewTitle">-</h3>
        <p id="previewMeta">-</p>
      </div>
      <div class="btn-group">
        <a id="btnDownload" href="#" class="btn-download">⬇ 다운로드</a>
        <button id="btnZip" class="btn-zip" onclick="downloadZip()">📦 규격 전체 ZIP</button>
      </div>
    </div>
    <div class="preview-body" id="previewBody">
      <div class="welcome-screen">
        <div style="font-size:56px;margin-bottom:14px">📄</div>
        <h2>문서를 선택하세요</h2>
        <p>(주)한영피엔에스 IMS 통합경영시스템<br>
        Word · Excel 문서를 브라우저에서 바로 미리볼 수 있습니다.<br><br>
        <span style="color:#ddd;font-size:12px">144개 문서 · 7개 ISO 규격 + IMS 통합</span></p>
      </div>
    </div>
  </div>
</div>

<!-- 토스트 알림 -->
<div class="toast" id="toast"></div>

<script>
// ─── 상태 ────────────────────────────────────────────
let currentStd = null;
let currentFile = null;
let allFiles = [];
let activeFilter = 'all';

// ─── 유틸 ────────────────────────────────────────────
function showToast(msg, duration=2200) {
  const t = document.getElementById('toast');
  t.textContent = msg;
  t.classList.add('show');
  setTimeout(() => t.classList.remove('show'), duration);
}

// ─── 초기화 ─────────────────────────────────────────
async function init() {
  await Promise.all([loadStats(), loadStandards()]);
}

async function loadStats() {
  try {
    const r = await fetch('/api/stats');
    const d = await r.json();
    document.getElementById('statTotal').textContent = d.total;
    document.getElementById('statWord').textContent = d.word;
    document.getElementById('statExcel').textContent = d.excel;
  } catch(e) {}
}

async function loadStandards() {
  try {
    const r = await fetch('/api/standards');
    const stds = await r.json();
    const container = document.getElementById('stdList');
    container.innerHTML = stds.map(s => `
      <div class="std-card" data-std="${s.key}" id="std-${s.key}">
        <div class="std-icon">${s.icon}</div>
        <div class="std-info">
          <div class="std-name">${s.name}</div>
          <div class="std-meta">문서 ${s.count}개</div>
        </div>
        <div class="std-badge" style="background:${s.color}">${s.badge}</div>
      </div>
    `).join('');
    // 이벤트 위임 방식 (특수문자 파일명 문제 방지)
    container.addEventListener('click', e => {
      const card = e.target.closest('.std-card');
      if (card) selectStandard(card.dataset.std);
    });
  } catch(e) { console.error('Standards load error:', e); }
}

// ─── 규격 선택 ──────────────────────────────────────
async function selectStandard(stdKey) {
  if (currentStd === stdKey) return;
  currentStd = stdKey;
  currentFile = null;
  activeFilter = 'all';

  document.querySelectorAll('.std-card').forEach(el => el.classList.remove('active'));
  const stdEl = document.getElementById('std-' + stdKey);
  if (stdEl) stdEl.classList.add('active');

  document.getElementById('previewHeaderBar').style.display = 'none';
  document.getElementById('previewBody').innerHTML = `
    <div class="welcome-screen">
      <div style="font-size:44px;margin-bottom:12px">👆</div>
      <h2>문서를 선택하세요</h2>
      <p>좌측 목록에서 문서를 클릭하면 미리보기가 표시됩니다</p>
    </div>`;

  const panel = document.getElementById('filePanel');
  panel.innerHTML = '<div class="loading"><div class="spinner"></div><div class="loading-text">파일 목록 로드 중...</div></div>';

  try {
    const r = await fetch('/api/files/' + stdKey);
    if (!r.ok) throw new Error('HTTP ' + r.status);
    allFiles = await r.json();
    renderFilePanel(allFiles, stdKey);
  } catch(e) {
    panel.innerHTML = '<div class="empty-state">❌ 파일 목록을 불러올 수 없습니다</div>';
    showToast('파일 목록 로드 실패');
  }
}

function renderFilePanel(files, stdKey) {
  const panel = document.getElementById('filePanel');
  const types = [...new Set(files.map(f => f.type))];

  const filteredFiles = activeFilter === 'all'
    ? files
    : files.filter(f => f.type === activeFilter);

  // 파일 아이템은 data-std, data-file 속성으로 → 이벤트 위임 처리
  panel.innerHTML = `
    <div class="file-panel-header">
      <div class="file-panel-title">${stdKey}</div>
      <div class="file-panel-count">전체 ${files.length}개 · 표시 ${filteredFiles.length}개</div>
      <div class="file-filter">
        <button class="filter-btn ${activeFilter==='all'?'active':''}" data-filter="all">전체</button>
        ${types.map(t => `<button class="filter-btn ${activeFilter===t?'active':''}" data-filter="${t}">${t}</button>`).join('')}
      </div>
    </div>
    <div class="file-list" id="fileListContent">
      ${filteredFiles.length === 0
        ? '<div class="empty-state">해당 유형의 문서가 없습니다</div>'
        : filteredFiles.map((f, idx) => {
            const safeDisplay = f.display.replace(/</g,'&lt;').replace(/>/g,'&gt;');
            const safeName = f.name.replace(/\\/g,'\\\\').replace(/'/g,"\\'");
            return `<div class="file-item" data-std="${stdKey}" data-file="${encodeURIComponent(f.name)}" data-idx="${idx}">
              <div class="file-ext ${f.ext.toLowerCase()}">${f.ext}</div>
              <div class="file-name">
                <strong title="${safeDisplay}">${safeDisplay}</strong>
                <small>${f.size} KB</small>
              </div>
              <div class="file-type-badge" style="background:${f.color}">${f.type}</div>
            </div>`;
          }).join('')}
    </div>`;

  // 이벤트 위임: 파일 클릭
  const listContent = document.getElementById('fileListContent');
  if (listContent) {
    listContent.addEventListener('click', e => {
      const item = e.target.closest('.file-item');
      if (item) {
        const std = item.dataset.std;
        const file = decodeURIComponent(item.dataset.file);
        previewFile(std, file, item);
      }
    });
  }

  // 이벤트 위임: 필터 버튼
  const filterBar = panel.querySelector('.file-filter');
  if (filterBar) {
    filterBar.addEventListener('click', e => {
      const btn = e.target.closest('.filter-btn');
      if (btn) setFilter(btn.dataset.filter);
    });
  }
}

function setFilter(type) {
  activeFilter = type;
  renderFilePanel(allFiles, currentStd);
}

// ─── 미리보기 ──────────────────────────────────────
async function previewFile(std, filename, itemEl) {
  if (currentFile === filename && currentStd === std) return;
  currentFile = filename;

  // 활성 파일 표시
  document.querySelectorAll('.file-item').forEach(el => el.classList.remove('active'));
  if (itemEl) itemEl.classList.add('active');

  const preview = document.getElementById('previewBody');
  preview.innerHTML = '<div class="loading"><div class="spinner"></div><div class="loading-text">미리보기 생성 중...</div></div>';

  // 헤더바 업데이트
  const displayName = filename.replace(/\.(docx|xlsx)$/i, '');
  const ext = filename.split('.').pop().toUpperCase();
  document.getElementById('previewTitle').textContent = displayName;
  document.getElementById('previewMeta').textContent = std + ' · ' + ext + ' 문서';
  document.getElementById('btnDownload').href = '/api/download/' + std + '/' + encodeURIComponent(filename);
  document.getElementById('previewHeaderBar').style.display = 'flex';

  try {
    const r = await fetch('/api/preview/' + std + '/' + encodeURIComponent(filename));
    if (!r.ok) throw new Error('HTTP ' + r.status);
    const d = await r.json();
    if (d.error) throw new Error(d.error);

    const extClass = ext === 'DOCX' ? 'docx' : 'xlsx';
    preview.innerHTML = `<div style="max-width:980px;margin:0 auto">${d.content}</div>`;
    preview.scrollTop = 0;
  } catch(e) {
    preview.innerHTML = `<div style="padding:20px"><div class="error-box">❌ 미리보기 오류: ${e.message}</div></div>`;
    showToast('미리보기 로드 실패: ' + e.message);
    console.error('Preview error:', e);
  }
}

// ─── ZIP 다운로드 ──────────────────────────────────
function downloadZip() {
  if (!currentStd) { showToast('먼저 규격을 선택하세요'); return; }
  showToast('📦 ZIP 파일 생성 중...');
  window.location.href = '/api/download-zip/' + currentStd;
}

// ─── 검색 ──────────────────────────────────────────
let searchTimer;
const searchInput = document.getElementById('searchInput');
const searchResults = document.getElementById('searchResults');

searchInput.addEventListener('input', function() {
  clearTimeout(searchTimer);
  const q = this.value.trim();
  if (q.length < 2) { searchResults.classList.remove('show'); return; }
  searchTimer = setTimeout(() => doSearch(q), 220);
});

// ✅ 수정: blur 대신 document click으로 처리 (검색 결과 클릭 전 닫히는 문제 방지)
document.addEventListener('click', function(e) {
  if (!e.target.closest('.header-search')) {
    searchResults.classList.remove('show');
  }
});

searchInput.addEventListener('focus', function() {
  if (this.value.trim().length >= 2) searchResults.classList.add('show');
});

async function doSearch(q) {
  try {
    const r = await fetch('/api/search?q=' + encodeURIComponent(q));
    const results = await r.json();

    if (results.length === 0) {
      searchResults.innerHTML = '<div class="search-item-empty">검색 결과가 없습니다</div>';
    } else {
      searchResults.innerHTML = results.map(item => {
        const safeDisplay = item.display.replace(/</g,'&lt;').replace(/>/g,'&gt;');
        return `<div class="search-item"
          data-std="${item.std}"
          data-file="${encodeURIComponent(item.name)}">
          <span class="s-type" style="background:${item.color}">${item.type}</span>
          <span style="flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${safeDisplay}</span>
          <span style="font-size:10px;color:#bbb;white-space:nowrap">${item.std}</span>
        </div>`;
      }).join('');

      // 이벤트 위임으로 검색 결과 클릭 처리
      searchResults.addEventListener('click', async function handler(e) {
        const si = e.target.closest('.search-item');
        if (!si) return;
        const std = si.dataset.std;
        const file = decodeURIComponent(si.dataset.file);
        searchResults.classList.remove('show');
        searchInput.value = '';
        await selectStandard(std);
        setTimeout(() => {
          // 파일 아이템 찾아서 클릭 시뮬레이션
          const items = document.querySelectorAll('.file-item');
          for (const item of items) {
            if (decodeURIComponent(item.dataset.file) === file) {
              item.click();
              item.scrollIntoView({ block: 'nearest', behavior: 'smooth' });
              break;
            }
          }
        }, 400);
        searchResults.removeEventListener('click', handler);
      }, { once: false });
    }
    searchResults.classList.add('show');
  } catch(e) { console.error('Search error:', e); }
}

// ─── Excel 시트 전환 ────────────────────────────────
function showSheet(idx) {
  document.querySelectorAll('.sheet-content').forEach((el, i) => {
    el.style.display = i === idx ? 'block' : 'none';
  });
  document.querySelectorAll('.sheet-tab').forEach((el, i) => {
    el.classList.toggle('active', i === idx);
  });
  // 탭 스크롤 위치 조정
  const tab = document.getElementById('tab-' + idx);
  if (tab) tab.scrollIntoView({ block: 'nearest', inline: 'nearest', behavior: 'smooth' });
}

// ─── 시작 ────────────────────────────────────────────
init();
</script>
</body>
</html>'''

# ─────────────────────────────────────────────────────
# 📷 모바일 바코드/QR 스캐너 페이지
# ─────────────────────────────────────────────────────
SCAN_TEMPLATE = r'''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
<title>📷 문서 바코드/QR 스캔 | 한영피엔에스 IMS</title>
<link rel="icon" type="image/svg+xml" href="/favicon.svg">
<script src="https://unpkg.com/html5-qrcode@2.3.8/html5-qrcode.min.js"></script>
<style>
* { box-sizing:border-box; margin:0; padding:0; -webkit-tap-highlight-color:transparent; }
:root { --brand:#1F3864; --brand-mid:#2E75B6; --brand-light:#D6E4F0; }
body { font-family:'Malgun Gothic','맑은 고딕',-apple-system,sans-serif;
       background:#f0f2f5; color:#222; padding-bottom:40px; }
.topbar {
  position:sticky; top:0; z-index:50; background:var(--brand); color:#fff;
  display:flex; align-items:center; gap:10px; padding:13px 16px;
  box-shadow:0 2px 8px rgba(0,0,0,.3);
}
.topbar a { color:#fff; text-decoration:none; font-size:22px; line-height:1; }
.topbar h1 { font-size:16px; font-weight:700; }
.topbar h1 span { display:block; font-size:10px; opacity:.7; font-weight:400; }
.wrap { max-width:560px; margin:0 auto; padding:16px; }

.scanner-card {
  background:#000; border-radius:14px; overflow:hidden; position:relative;
  box-shadow:0 4px 18px rgba(0,0,0,.25); margin-bottom:14px;
}
#reader { width:100%; min-height:60px; }
#reader video { width:100% !important; display:block; }
.scanner-hint {
  text-align:center; color:#cfe0f5; font-size:13px; padding:12px;
}
.btn {
  display:flex; align-items:center; justify-content:center; gap:8px; width:100%;
  border:none; border-radius:11px; padding:15px; font-size:16px; font-weight:700;
  cursor:pointer; transition:.15s;
}
.btn-primary { background:var(--brand); color:#fff; }
.btn-primary:active { background:#16294a; }
.btn-stop { background:#b02a2a; color:#fff; }
.btn-ghost { background:#fff; color:var(--brand); border:1.5px solid var(--brand-light); }

.manual { display:flex; gap:8px; margin:14px 0; }
.manual input {
  flex:1; padding:13px 14px; border:1.5px solid #d0d7e2; border-radius:11px;
  font-size:15px; outline:none;
}
.manual input:focus { border-color:var(--brand-mid); }
.manual button { flex-shrink:0; width:auto; padding:13px 18px; }

.status {
  text-align:center; font-size:13px; color:#667; margin:6px 0 14px;
  min-height:18px;
}
.status.err { color:#b02a2a; font-weight:600; }

/* ─ 결과 ─ */
.result-head {
  display:flex; align-items:center; gap:10px; margin:18px 0 10px;
  padding-bottom:8px; border-bottom:2px solid var(--brand-light);
}
.result-head .scanned-code {
  font-family:monospace; background:var(--brand-light); color:var(--brand);
  padding:4px 9px; border-radius:6px; font-size:14px; font-weight:700;
}
.result-head small { color:#889; font-size:12px; }

.card {
  background:#fff; border-radius:12px; padding:14px; margin-bottom:11px;
  box-shadow:0 2px 8px rgba(0,0,0,.07); border-left:4px solid #ccc;
}
.card .row1 { display:flex; align-items:center; gap:8px; flex-wrap:wrap; margin-bottom:4px; }
.badge { font-size:11px; padding:3px 8px; border-radius:5px; color:#fff; font-weight:700; }
.ext { font-size:10px; padding:2px 6px; border-radius:4px; background:#eef1f6; color:#556; font-weight:700; }
.card .title { font-size:15px; font-weight:700; color:#1a1a1a; line-height:1.35; margin:2px 0; }
.card .meta { font-size:12px; color:#7a8594; }
.card .actions { display:flex; gap:8px; margin-top:11px; flex-wrap:wrap; }
.card .actions a, .card .actions button {
  flex:1; min-width:90px; text-align:center; text-decoration:none;
  padding:10px; border-radius:9px; font-size:13px; font-weight:700;
  border:none; cursor:pointer;
}
.act-view { background:var(--brand); color:#fff; }
.act-dl   { background:#eef3fb; color:var(--brand); }
.act-qr   { background:#f3eefb; color:#6b2fae; }

.empty { text-align:center; padding:40px 16px; color:#9aa3af; }
.empty .big { font-size:46px; margin-bottom:10px; }

/* 미리보기 모달 */
.modal {
  display:none; position:fixed; inset:0; z-index:200;
  background:rgba(0,0,0,.55); padding:0;
}
.modal.show { display:block; }
.modal-inner {
  position:absolute; inset:0; background:#fff; display:flex; flex-direction:column;
}
.modal-head {
  display:flex; align-items:center; gap:10px; padding:13px 16px;
  background:var(--brand); color:#fff; flex-shrink:0;
}
.modal-head .t { flex:1; font-size:14px; font-weight:700;
  overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.modal-head button { background:rgba(255,255,255,.2); color:#fff; border:none;
  border-radius:8px; padding:8px 13px; font-size:14px; font-weight:700; cursor:pointer; }
.modal-body { flex:1; overflow:auto; padding:16px; -webkit-overflow-scrolling:touch; }
.modal-body img.qr { display:block; margin:20px auto; width:230px; height:230px; }
.modal-body .qr-cap { text-align:center; color:#667; font-size:13px; line-height:1.6; }
.modal-body .qr-cap code { background:#eef1f6; padding:2px 7px; border-radius:5px; }

/* 문서 미리보기 내용 스타일 */
.doc-content { font-size:14px; line-height:1.7; color:#222; }
.doc-content h1,.doc-content h2,.doc-content h3 { margin:14px 0 8px; color:var(--brand); }
.doc-content p { margin:6px 0; }
.doc-content table { border-collapse:collapse; width:100%; margin:12px 0; font-size:12px; }
.doc-content th,.doc-content td { border:1px solid #ccc; padding:6px 8px; text-align:left; }
.doc-content th { background:var(--brand-light); }
.error-box { color:#b02a2a; padding:20px; text-align:center; }
.spinner { text-align:center; padding:40px; color:#889; }
</style>
</head>
<body>
<div class="topbar">
  <a href="/" title="홈으로">←</a>
  <h1>문서 바코드 · QR 스캔<span>(주)한영피엔에스 IMS</span></h1>
</div>

<div class="wrap">
  <div class="scanner-card">
    <div id="reader"></div>
    <div class="scanner-hint" id="hint">아래 버튼을 눌러 카메라로 문서 라벨의 QR·바코드를 비춰주세요.</div>
  </div>

  <button class="btn btn-primary" id="startBtn">📷 카메라 스캔 시작</button>
  <button class="btn btn-stop" id="stopBtn" style="display:none">■ 스캔 중지</button>

  <div class="manual">
    <input type="text" id="manualInput" placeholder="문서코드 직접 입력 (예: QMS-G-001)" autocomplete="off">
    <button class="btn btn-primary" id="manualBtn">조회</button>
  </div>

  <div class="status" id="status"></div>
  <div id="results"></div>
</div>

<!-- 미리보기 / QR 모달 -->
<div class="modal" id="modal">
  <div class="modal-inner">
    <div class="modal-head">
      <div class="t" id="modalTitle"></div>
      <button id="modalClose">닫기 ✕</button>
    </div>
    <div class="modal-body" id="modalBody"></div>
  </div>
</div>

<script>
const $ = s => document.querySelector(s);
const resultsEl = $('#results'), statusEl = $('#status'), hintEl = $('#hint');
let html5Qr = null, scanning = false, lastCode = '', lastTs = 0;

function setStatus(msg, isErr){
  statusEl.textContent = msg || '';
  statusEl.className = 'status' + (isErr ? ' err' : '');
}

function escapeHtml(s){ return (s||'').replace(/[&<>"']/g, c =>
  ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }

async function lookup(code){
  if(!code) return;
  setStatus('🔎 "' + code + '" 조회 중...');
  resultsEl.innerHTML = '<div class="spinner">⏳ 조회 중...</div>';
  try{
    const r = await fetch('/api/lookup?code=' + encodeURIComponent(code));
    const data = await r.json();
    render(data);
  }catch(e){
    setStatus('조회 실패: ' + e.message, true);
    resultsEl.innerHTML = '';
  }
}

function render(data){
  const items = data.results || [];
  setStatus('');
  let head = '<div class="result-head"><span class="scanned-code">' +
    escapeHtml(data.query) + '</span><small>' + items.length + '건 조회됨</small></div>';

  if(data.kind === 'none' || items.length === 0){
    resultsEl.innerHTML = head +
      '<div class="empty"><div class="big">🔍</div>' +
      '<b>일치하는 문서가 없습니다.</b><br>' +
      '문서코드(예: QMS-G-001)나 규격(예: ISO9001)을 확인해 주세요.</div>';
    return;
  }

  if(data.kind === 'standard' && data.standard){
    const s = data.standard;
    head = '<div class="result-head"><span class="scanned-code" style="background:' +
      s.color + ';color:#fff">' + s.icon + ' ' + escapeHtml(s.name) + '</span>' +
      '<small>' + items.length + '개 문서</small></div>';
  }

  const cards = items.map(d => {
    const url = '/api/preview/' + encodeURIComponent(d.std) + '/' + encodeURIComponent(d.name);
    const dl  = '/api/download/' + encodeURIComponent(d.std) + '/' + encodeURIComponent(d.name);
    return '<div class="card" style="border-left-color:' + d.color + '">' +
      '<div class="row1">' +
        '<span class="badge" style="background:' + d.color + '">' + escapeHtml(d.type) + '</span>' +
        '<span class="ext">' + escapeHtml(d.ext) + '</span>' +
        '<span class="meta">' + escapeHtml(d.code) + ' · ' + escapeHtml(d.std_name) + '</span>' +
      '</div>' +
      '<div class="title">' + escapeHtml(d.display) + '</div>' +
      '<div class="actions">' +
        '<button class="act-view" onclick="preview(\'' + d.std + '\',\'' +
            encodeURIComponent(d.name) + '\',\'' + escapeHtml(d.display).replace(/'/g,"\\'") + '\')">📄 미리보기</button>' +
        '<a class="act-dl" href="' + dl + '">⬇ 다운로드</a>' +
        '<button class="act-qr" onclick="showQr(\'' + escapeHtml(d.code) + '\',\'' +
            escapeHtml(d.display).replace(/'/g,"\\'") + '\')">🔳 QR</button>' +
      '</div></div>';
  }).join('');

  resultsEl.innerHTML = head + cards;
}

async function preview(std, encName, title){
  openModal(title);
  $('#modalBody').innerHTML = '<div class="spinner">⏳ 문서 불러오는 중...</div>';
  try{
    const r = await fetch('/api/preview/' + encodeURIComponent(std) + '/' + encName);
    const data = await r.json();
    $('#modalBody').innerHTML = data.content ||
      '<div class="error-box">' + (data.error || '미리보기를 불러올 수 없습니다') + '</div>';
  }catch(e){
    $('#modalBody').innerHTML = '<div class="error-box">불러오기 실패: ' + e.message + '</div>';
  }
}

function showQr(code, title){
  openModal('QR 코드 · ' + title);
  // 휴대폰 기본 카메라로 스캔하면 이 페이지가 열리도록 전체 URL을 인코딩
  const link = location.origin + '/scan?code=' + encodeURIComponent(code);
  const qrSrc = '/api/qr?data=' + encodeURIComponent(link);
  $('#modalBody').innerHTML =
    '<img class="qr" src="' + qrSrc + '" alt="QR ' + escapeHtml(code) + '">' +
    '<div class="qr-cap">문서코드 <code>' + escapeHtml(code) + '</code><br>' +
    '이 QR을 인쇄해 문서·설비에 부착하면<br>휴대폰으로 스캔하여 바로 조회할 수 있습니다.</div>';
}

function openModal(title){
  $('#modalTitle').textContent = title || '';
  $('#modal').classList.add('show');
}
$('#modalClose').onclick = () => $('#modal').classList.remove('show');

// ─ 카메라 스캔 ─
function onScanSuccess(decodedText){
  const now = Date.now();
  if(decodedText === lastCode && now - lastTs < 2500) return; // 중복 방지
  lastCode = decodedText; lastTs = now;
  if(navigator.vibrate) navigator.vibrate(120);
  lookup(decodedText);
}

async function startScan(){
  if(scanning) return;
  if(typeof Html5Qrcode === 'undefined'){
    setStatus('스캐너 라이브러리를 불러오지 못했습니다. 인터넷 연결을 확인하세요.', true);
    return;
  }
  html5Qr = new Html5Qrcode('reader', { verbose:false });
  const config = {
    fps: 10,
    qrbox: { width: 250, height: 250 },
    formatsToSupport: undefined  // QR + 주요 1D 바코드 모두 지원
  };
  try{
    await html5Qr.start({ facingMode: 'environment' }, config, onScanSuccess, ()=>{});
    scanning = true;
    hintEl.style.display = 'none';
    $('#startBtn').style.display = 'none';
    $('#stopBtn').style.display = 'flex';
    setStatus('📷 코드를 화면 안에 비춰주세요...');
  }catch(e){
    setStatus('카메라를 시작할 수 없습니다: ' + e + ' (HTTPS 환경 및 카메라 권한 필요)', true);
  }
}

async function stopScan(){
  if(html5Qr && scanning){
    try{ await html5Qr.stop(); html5Qr.clear(); }catch(e){}
  }
  scanning = false;
  hintEl.style.display = 'block';
  $('#startBtn').style.display = 'flex';
  $('#stopBtn').style.display = 'none';
  setStatus('');
}

$('#startBtn').onclick = startScan;
$('#stopBtn').onclick = stopScan;
$('#manualBtn').onclick = () => lookup($('#manualInput').value.trim());
$('#manualInput').addEventListener('keydown', e => { if(e.key==='Enter') lookup(e.target.value.trim()); });

// URL ?code= 로 진입(휴대폰 기본 카메라가 QR URL을 연 경우) → 자동 조회
(function(){
  const p = new URLSearchParams(location.search);
  const c = p.get('code') || p.get('q');
  if(c){ $('#manualInput').value = c; lookup(c); }
})();
</script>
</body>
</html>'''


@app.route('/scan')
def scan_page():
    return render_template_string(SCAN_TEMPLATE)


@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 3000))
    print(f'🚀 IMS 문서 미리보기 서버 v2.2 시작')
    print(f'   URL: http://0.0.0.0:{port}')
    print(f'   문서 경로: {DOCS_ROOT}')
    app.run(host='0.0.0.0', port=port, debug=False)
